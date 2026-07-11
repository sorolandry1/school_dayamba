import io
import zipfile
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from rest_framework import status
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser
from django.http import HttpResponse
from django.db.models import Sum, Count, Q, Avg as DAvg
from apps.users.permissions import IsAdmin, IsAdminOrTeacher, CanViewReports, IsCashManager, IsPlatformAdmin
from apps.students.models import Student
from apps.grades.models import Grade
from apps.subjects.models import Subject
from apps.attendance.models import Attendance
from apps.payments.models import Payment
from apps.classes.models import Classe
from utils.pdf_generator import (
    generate_bulletin_pdf, generate_bulletin_pdf_templated,
    generate_receipt_pdf_templated, generate_card_pdf_templated,
    generate_class_list_pdf, generate_subject_sheet_pdf,
    generate_student_sheet_pdf, generate_class_roster_pdf, generate_timetable_pdf,
    generate_teacher_timetable_pdf,
)
from django.db.models import Avg, Count, Q, F, FloatField
from django.db.models.functions import Cast


def _normalized_avg_expr():
    """Expression ORM de la moyenne normalisée sur 20 : ``value * 20 / max_value``.

    Ramène chaque note sur /20 AVANT de moyenner, au lieu de moyenner des valeurs
    brutes d'échelles hétérogènes (une note /10 ou /100 fausserait ``Avg('value')``).
    Miroir SQL de ``Grade.normalized_value``.
    """
    return Avg(Cast('value', FloatField()) * 20.0 / Cast('max_value', FloatField()))


def _compute_rankings_map(students, subjects, period):
    """Classement d'une classe calculé en UNE requête agrégée (évite le N+1).

    Retourne ``{student_id: {rank, total_students, general_average, class_average}}``.
    Chaque moyenne par matière est normalisée sur 20 puis pondérée par le
    coefficient de la matière. Seules les matières où l'élève a au moins une note
    contribuent à la pondération (comportement identique à l'ancien calcul).
    """
    students = list(students)
    student_ids = [s.id for s in students]
    subject_coeff = {subj.id: float(subj.coefficient) for subj in subjects}
    if not student_ids:
        return {}

    # Moyenne normalisée par (élève, matière) en une seule requête groupée.
    per_pair = {}
    if subject_coeff:
        grade_q = Grade.objects.filter(
            student_id__in=student_ids, subject_id__in=subject_coeff.keys()
        )
        if period:
            grade_q = grade_q.filter(period=period)
        for row in grade_q.values('student_id', 'subject_id').annotate(avg=_normalized_avg_expr()):
            if row['avg'] is not None:
                per_pair[(row['student_id'], row['subject_id'])] = float(row['avg'])

    averages = []
    for sid in student_ids:
        total_weighted = 0.0
        total_coeff = 0.0
        for subj_id, coeff in subject_coeff.items():
            avg = per_pair.get((sid, subj_id))
            if avg is not None:
                total_weighted += avg * coeff
                total_coeff += coeff
        general_avg = round(total_weighted / total_coeff, 2) if total_coeff > 0 else 0
        averages.append({'student_id': sid, 'average': general_avg})

    averages.sort(key=lambda x: x['average'], reverse=True)
    total_students = len(averages)
    class_avg = round(sum(a['average'] for a in averages) / total_students, 2) if averages else 0
    return {
        entry['student_id']: {
            'rank': i + 1,
            'total_students': total_students,
            'general_average': entry['average'],
            'class_average': class_avg,
        }
        for i, entry in enumerate(averages)
    }


def _selected_ecole(request):
    from apps.users.models import Ecole
    u = getattr(request, 'user', None)
    if u and u.is_authenticated and u.role != 'ADMIN':
        return getattr(u, 'ecole', None)
    eid = request.query_params.get('ecole')
    if eid and str(eid).isdigit():
        return Ecole.objects.filter(id=int(eid)).first()
    return None


def _ecole_id(request):
    """École à utiliser pour filtrer les agrégats : celle de l'utilisateur (non
    ADMIN), ou `?ecole=` pour le super-admin, sinon None (toutes écoles)."""
    selected = _selected_ecole(request)
    return getattr(selected, 'id', None)


def _teacher_class_ids(user):
    """Ensemble des ids de classes autorisées pour un professeur : ses classes
    assignées + les classes où il enseigne une matière."""
    teacher = getattr(user, 'teacher_profile', None)
    if teacher is None:
        return set()
    ids = set(teacher.classes.values_list('id', flat=True))
    ids |= set(Classe.objects.filter(subjects__teacher=teacher).values_list('id', flat=True))
    return ids


def _teacher_can_access_classe(user, classe_id):
    """Un professeur ne peut consulter que ses classes assignées ; les autres
    rôles (admin, directeur, éducateur) ne sont pas restreints ici."""
    if getattr(user, 'role', None) != 'TEACHER':
        return True
    return int(classe_id) in _teacher_class_ids(user)


def _bulletin_extra(student, rankings_data, period_raw):
    """Construit le dict `extra` (période, assiduité, distinction, décision)
    passé au générateur de bulletin PDF."""
    from apps.reports.models import PlatformSettings
    from apps.reports.academics import (
        normalize_period, period_label, period_date_range,
        attendance_hours, year_end_decision, distinction,
    )
    ps = PlatformSettings.get_solo(ecole=getattr(student, 'ecole', None))
    system = ps.period_system
    period = normalize_period(period_raw)
    is_annual = period is None
    ay = student.classe.academic_year if student.classe else None
    start, end = period_date_range(ay, period_raw, system)
    avg = (rankings_data or {}).get('general_average')
    return {
        'period': period,
        'period_label': period_label(period_raw, system),
        'attendance': attendance_hours(student, start, end, ps.hours_per_day),
        'distinction': distinction(avg, ps),
        'is_annual': is_annual,
        'decision': year_end_decision(avg, ps) if is_annual else None,
    }


class BulletinPDFView(APIView):
    permission_classes = [CanViewReports]

    def _get_template_config(self, request):
        return _resolve_template(
            'bulletin',
            template_id=request.query_params.get('template_id'),
            school_type=request.query_params.get('school_type'),
            request=request,
        )

    def get(self, request, student_id):
        eid = _ecole_id(request)
        students = Student.objects.select_related('classe', 'classe__level').filter(id=student_id)
        if eid:
            students = students.filter(ecole_id=eid)
        student = students.first()
        if not student:
            return Response({'error': 'Élève non trouvé.'}, status=404)

        period_raw = request.query_params.get('period')
        subjects = Subject.objects.select_related('teacher', 'teacher__user').filter(classe=student.classe)
        rankings_data = self._calculate_rankings(student, period_raw)
        template_config = self._get_template_config(request)
        extra = _bulletin_extra(student, rankings_data, period_raw)

        # Toujours via le générateur templaté (comme les cartes et les reçus) :
        # le design est personnalisable par l'administrateur ; sinon, valeurs
        # par défaut + identité de l'établissement (Super-Admin).
        pdf_buffer = generate_bulletin_pdf_templated(
            student, subjects, rankings_data, template_config or {}, extra=extra
        )

        response = HttpResponse(pdf_buffer, content_type='application/pdf')
        suffix = f"_{extra['period']}" if extra.get('period') else '_annuel'
        response['Content-Disposition'] = (
            f'attachment; filename="bulletin_{student.matricule}{suffix}.pdf"'
        )
        return response

    def _calculate_rankings(self, student, period_raw=None):
        if not student.classe:
            return {}
        from apps.reports.academics import normalize_period
        period = normalize_period(period_raw)
        subjects = Subject.objects.filter(classe=student.classe)
        students_in_class = Student.objects.filter(classe=student.classe, is_active=True)
        rankings_map = _compute_rankings_map(students_in_class, subjects, period)
        return rankings_map.get(student.id, {})


class BulletinClassePDFView(APIView):
    """Generate a ZIP containing all bulletins for a given class."""
    permission_classes = [CanViewReports]

    def get(self, request, classe_id):
        eid = _ecole_id(request)
        classes = Classe.objects.select_related('academic_year', 'level').filter(id=classe_id)
        if eid:
            classes = classes.filter(ecole_id=eid)
        classe = classes.first()
        if not classe:
            return Response({'error': 'Classe non trouvée.'}, status=404)

        students = Student.objects.filter(classe=classe, is_active=True).order_by('last_name', 'first_name')
        if not students.exists():
            return Response({'error': 'Aucun élève dans cette classe.'}, status=404)

        subjects = Subject.objects.select_related('teacher', 'teacher__user').filter(classe=classe)

        # Template support (modèle de l'éditeur ou défaut par type d'établissement)
        template_config = _resolve_template(
            'bulletin',
            template_id=request.query_params.get('template_id'),
            school_type=request.query_params.get('school_type'),
            request=request,
        )

        period_raw = request.query_params.get('period')
        # Pre-calculate rankings for all students once
        rankings_map = self._calculate_all_rankings(students, subjects, period_raw)

        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
            for student in students:
                rankings_data = rankings_map.get(student.id, {})
                extra = _bulletin_extra(student, rankings_data, period_raw)
                pdf_buffer = generate_bulletin_pdf_templated(
                    student, subjects, rankings_data, template_config or {}, extra=extra
                )
                filename = f"bulletin_{student.matricule}_{student.last_name}_{student.first_name}.pdf"
                zf.writestr(filename, pdf_buffer.read())

        zip_buffer.seek(0)
        from apps.reports.academics import normalize_period
        psuffix = normalize_period(period_raw)
        safe_name = classe.name.replace(' ', '_').replace('/', '-')
        safe_name += f"_p{psuffix}" if psuffix else '_annuel'
        response = HttpResponse(zip_buffer, content_type='application/zip')
        response['Content-Disposition'] = f'attachment; filename="bulletins_{safe_name}.zip"'
        return response

    def _calculate_all_rankings(self, students, subjects, period_raw=None):
        from apps.reports.academics import normalize_period
        period = normalize_period(period_raw)
        return _compute_rankings_map(students, subjects, period)


class DashboardStatsView(APIView):
    permission_classes = [IsAdmin]
    CACHE_TTL = 60  # seconds — short enough for real-time attendance to feel live

    def get(self, request):
        from django.core.cache import cache
        eid = _ecole_id(request)
        cache_key = f'dashboard_stats_{eid or "all"}'
        cached = cache.get(cache_key)
        if cached is not None:
            return Response(cached)

        students_qs = Student.objects.filter(is_active=True)
        payments_qs = Payment.objects.all()
        attendance_qs = Attendance.objects.all()
        from apps.payments.models import Expense
        expenses_qs = Expense.objects.all()
        from apps.classes.models import Classe, AcademicYear
        from apps.teachers.models import Teacher
        classes_qs = Classe.objects.select_related('level')
        years_qs = AcademicYear.objects.filter(is_current=True)
        teachers_qs = Teacher.objects.all()
        if eid:
            students_qs = students_qs.filter(ecole_id=eid)
            payments_qs = payments_qs.filter(ecole_id=eid)
            attendance_qs = attendance_qs.filter(student__ecole_id=eid)
            expenses_qs = expenses_qs.filter(ecole_id=eid)
            classes_qs = classes_qs.filter(ecole_id=eid)
            years_qs = years_qs.filter(ecole_id=eid)
            teachers_qs = teachers_qs.filter(ecole_id=eid)

        total_students = students_qs.count()
        total_teachers = teachers_qs.count()
        current_year = years_qs.first()
        total_classes = classes_qs.filter(academic_year=current_year).count() if current_year else 0

        from django.db.models import Sum as _Sum
        payment_stats = {
            'total_collected': payments_qs.filter(status='PAID').aggregate(t=_Sum('amount'))['t'] or 0,
            'total_pending': payments_qs.filter(status='PENDING').aggregate(t=_Sum('amount'))['t'] or 0,
            'total_overdue': payments_qs.filter(status='OVERDUE').aggregate(t=_Sum('amount'))['t'] or 0,
            'count_paid': payments_qs.filter(status='PAID').count(),
            'count_pending': payments_qs.filter(status='PENDING').count(),
            'count_overdue': payments_qs.filter(status='OVERDUE').count(),
        }

        from django.utils import timezone
        today = timezone.now().date()
        today_attendance = attendance_qs.filter(date=today)
        scanned_in = today_attendance.filter(status__in=['PRESENT', 'LATE']).count()
        attendance_stats = {
            'present_today': today_attendance.filter(status='PRESENT').count(),
            'late_today': today_attendance.filter(status='LATE').count(),
            'absent_today': max(0, total_students - scanned_in),
        }

        expense_stats = {
            'total_expenses': float(expenses_qs.aggregate(t=_Sum('amount'))['t'] or 0),
            'net_balance': float(
                (payments_qs.filter(status='PAID').aggregate(t=_Sum('amount'))['t'] or 0) -
                (expenses_qs.aggregate(t=_Sum('amount'))['t'] or 0)
            ),
        }

        result = {
            'total_students': total_students,
            'total_teachers': total_teachers,
            'total_classes': total_classes,
            'payments': payment_stats,
            'attendance_today': attendance_stats,
            'expenses': expense_stats,
        }
        cache.set(cache_key, result, self.CACHE_TTL)
        return Response(result)


class AccountingView(APIView):
    """Comptabilité : compte de résultat, caisse, banque, bilan simplifié."""
    permission_classes = [IsCashManager]

    def get(self, request):
        from django.db.models import Sum
        from apps.payments.models import Expense

        eid = _ecole_id(request)
        payments_qs = Payment.objects.all()
        expenses_qs = Expense.objects.all()
        if eid:
            payments_qs = payments_qs.filter(ecole_id=eid)
            expenses_qs = expenses_qs.filter(ecole_id=eid)

        paid = payments_qs.filter(status='PAID')
        produits = float(paid.aggregate(t=Sum('amount'))['t'] or 0)
        charges = float(expenses_qs.aggregate(t=Sum('amount'))['t'] or 0)
        charges_salaires = float(
            expenses_qs.filter(category='SALAIRES').aggregate(t=Sum('amount'))['t'] or 0
        )
        charges_autres = charges - charges_salaires
        resultat = produits - charges
        creances = float(payments_qs.filter(status__in=['PENDING', 'OVERDUE'])
                         .aggregate(t=Sum('amount'))['t'] or 0)

        def solde(method):
            r = float(paid.filter(method=method).aggregate(t=Sum('amount'))['t'] or 0)
            d = float(expenses_qs.filter(method=method).aggregate(t=Sum('amount'))['t'] or 0)
            return {'recettes': r, 'depenses': d, 'solde': r - d}

        caisse = solde('CAISSE')
        banque = solde('BANQUE')
        tresorerie = caisse['solde'] + banque['solde']

        return Response({
            'compte_resultat': {
                'produits': produits, 'charges': charges,
                'charges_salaires': charges_salaires, 'charges_autres': charges_autres,
                'resultat': resultat,
            },
            'caisse': caisse,
            'banque': banque,
            'bilan': {
                'tresorerie': tresorerie,
                'creances': creances,
                'total_actif': tresorerie + creances,
                'resultat': resultat,
            },
        })


class AccountingExportView(APIView):
    """Export comptable Excel de toutes les écritures (recettes + dépenses)."""
    permission_classes = [IsCashManager]

    def get(self, request):
        from apps.payments.models import Expense
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Journal comptable'
        headers = ['Date', 'Type', 'Libellé', 'Méthode', 'Recette (FCFA)', 'Dépense (FCFA)']
        ws.append(headers)
        for col in range(1, len(headers) + 1):
            c = ws.cell(row=1, column=col)
            c.font = Font(bold=True, color='FFFFFF')
            c.fill = PatternFill('solid', fgColor='1E40AF')
            c.alignment = Alignment(horizontal='center')

        eid = _ecole_id(request)
        pay_qs = Payment.objects.filter(status='PAID').select_related('student')
        exp_qs = Expense.objects.all()
        if eid:
            pay_qs = pay_qs.filter(ecole_id=eid)
            exp_qs = exp_qs.filter(ecole_id=eid)
        rows = []
        for p in pay_qs:
            rows.append((p.payment_date, 'Recette',
                         f"{p.get_payment_type_display()} — {p.student.full_name}",
                         dict(Payment._meta.get_field('method').choices).get(p.method, p.method),
                         float(p.amount), 0))
        for e in exp_qs:
            rows.append((e.date, 'Dépense', f"{e.label} ({e.get_category_display()})",
                         e.get_method_display(), 0, float(e.amount)))
        rows.sort(key=lambda r: str(r[0]))

        tot_r = tot_d = 0
        for r in rows:
            ws.append([str(r[0]), r[1], r[2], r[3], r[4] or '', r[5] or ''])
            tot_r += r[4]; tot_d += r[5]
        ws.append([])
        ws.append(['', '', 'TOTAUX', '', tot_r, tot_d])
        ws.append(['', '', 'RÉSULTAT', '', tot_r - tot_d, ''])
        for i, w in enumerate([14, 12, 40, 12, 16, 16], 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        resp = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        resp['Content-Disposition'] = 'attachment; filename="comptabilite.xlsx"'
        return resp


class ChartsView(APIView):
    """Séries pour les graphiques : évolution paiements, finances, absentéisme, réussite."""
    permission_classes = [IsAdmin]

    def get(self, request):
        from django.db.models.functions import TruncMonth
        from django.db.models import Sum, Count, Q, Avg
        from apps.payments.models import Expense
        from apps.classes.models import Classe

        def fmt(d):
            return d.strftime('%m/%Y') if d else '—'

        eid = _ecole_id(request)
        payments_base = Payment.objects.all()
        expenses_base = Expense.objects.all()
        attendance_base = Attendance.objects.all()
        classes_base = Classe.objects.select_related('level').all()
        if eid:
            payments_base = payments_base.filter(ecole_id=eid)
            expenses_base = expenses_base.filter(ecole_id=eid)
            attendance_base = attendance_base.filter(student__ecole_id=eid)
            classes_base = classes_base.filter(ecole_id=eid)

        # Paiements encaissés par mois
        pay = (payments_base.filter(status='PAID')
               .annotate(m=TruncMonth('payment_date')).values('m')
               .annotate(total=Sum('amount')).order_by('m'))
        payments_evolution = [{'period': fmt(r['m']), 'montant': float(r['total'] or 0)} for r in pay]

        # Finances : recettes vs dépenses par mois
        exp = (expenses_base.annotate(m=TruncMonth('date')).values('m')
               .annotate(total=Sum('amount')).order_by('m'))
        months = {}
        for r in pay:
            months.setdefault(r['m'], {'recettes': 0.0, 'depenses': 0.0})['recettes'] = float(r['total'] or 0)
        for r in exp:
            months.setdefault(r['m'], {'recettes': 0.0, 'depenses': 0.0})['depenses'] = float(r['total'] or 0)
        finance = [
            {'period': fmt(m), 'recettes': v['recettes'], 'depenses': v['depenses'],
             'solde': v['recettes'] - v['depenses']}
            for m, v in sorted(months.items(), key=lambda x: (x[0] is None, x[0]))
        ]

        # Absentéisme par mois
        att = (attendance_base.annotate(m=TruncMonth('date')).values('m')
               .annotate(total=Count('id'),
                         absent=Count('id', filter=Q(status='ABSENT')),
                         late=Count('id', filter=Q(status='LATE'))).order_by('m'))
        absenteeism = [
            {'period': fmt(r['m']),
             'taux_absence': round(r['absent'] / r['total'] * 100, 1) if r['total'] else 0,
             'taux_retard': round(r['late'] / r['total'] * 100, 1) if r['total'] else 0}
            for r in att
        ]

        # Réussite par classe (moyenne par élève → moyenne de classe + taux ≥ 10)
        success = []
        for c in classes_base:
            ps = list(Grade.objects.filter(student__classe=c).values('student').annotate(avg=Avg('value')))
            if not ps:
                continue
            passed = sum(1 for x in ps if (x['avg'] or 0) >= 10)
            success.append({
                'classe': c.name,
                'moyenne': round(sum(float(x['avg'] or 0) for x in ps) / len(ps), 2),
                'taux_reussite': round(passed / len(ps) * 100, 1),
            })

        return Response({
            'payments_evolution': payments_evolution,
            'finance': finance,
            'absenteeism': absenteeism,
            'success': success,
        })


class AttendanceReportView(APIView):
    """Absence report per class with optional date range."""
    permission_classes = [CanViewReports]

    def get(self, request):
        classe_id = request.query_params.get('classe_id')
        date_from = request.query_params.get('date_from')
        date_to = request.query_params.get('date_to')

        eid = _ecole_id(request)
        students_qs = Student.objects.filter(is_active=True)
        attendance_qs = Attendance.objects.all()
        if eid:
            students_qs = students_qs.filter(ecole_id=eid)
            attendance_qs = attendance_qs.filter(student__ecole_id=eid)
        if classe_id:
            students_qs = students_qs.filter(classe_id=classe_id)
            attendance_qs = attendance_qs.filter(student__classe_id=classe_id)
        if date_from:
            attendance_qs = attendance_qs.filter(date__gte=date_from)
        if date_to:
            attendance_qs = attendance_qs.filter(date__lte=date_to)

        # Per-student summary
        result = []
        for student in students_qs.select_related('classe'):
            records = attendance_qs.filter(student=student)
            agg = records.aggregate(
                total=Count('id'),
                present=Count('id', filter=Q(status='PRESENT')),
                absent=Count('id', filter=Q(status='ABSENT')),
                late=Count('id', filter=Q(status='LATE')),
            )
            total_hours_absent = sum(
                r.hours_absent for r in records if r.hours_absent
            )
            result.append({
                'student_id': student.id,
                'student_name': student.full_name,
                'classe': student.classe.name if student.classe else '—',
                'total_days': agg['total'],
                'present': agg['present'],
                'absent': agg['absent'],
                'late': agg['late'],
                'hours_absent': round(total_hours_absent, 1),
            })

        result.sort(key=lambda x: (x['classe'], x['student_name']))

        # Class-level aggregates
        global_stats = attendance_qs.aggregate(
            total=Count('id'),
            present=Count('id', filter=Q(status='PRESENT')),
            absent=Count('id', filter=Q(status='ABSENT')),
            late=Count('id', filter=Q(status='LATE')),
        )

        return Response({'students': result, 'global': global_stats})


class AttendanceReportPDFView(APIView):
    """Generate a PDF attendance report for daily/weekly/monthly or custom ranges."""
    permission_classes = [CanViewReports]

    def get(self, request):
        classe_id = request.query_params.get('classe_id')
        period = request.query_params.get('period')
        date_from = request.query_params.get('date_from')
        date_to = request.query_params.get('date_to')

        eid = _ecole_id(request)
        classe_name = None
        if classe_id:
            classe_qs = Classe.objects.filter(id=classe_id)
            if eid:
                classe_qs = classe_qs.filter(ecole_id=eid)
            classe = classe_qs.first()
            if not classe:
                return Response({'error': 'Classe introuvable.'}, status=404)
            classe_name = classe.name

        students_qs = Student.objects.filter(is_active=True)
        attendance_qs = Attendance.objects.all()
        if eid:
            students_qs = students_qs.filter(ecole_id=eid)
            attendance_qs = attendance_qs.filter(student__ecole_id=eid)
        if classe_id:
            students_qs = students_qs.filter(classe_id=classe_id)
            attendance_qs = attendance_qs.filter(student__classe_id=classe_id)

        start_date, end_date = self._resolve_date_range(period, date_from, date_to)
        attendance_qs = attendance_qs.filter(date__gte=start_date, date__lte=end_date)

        result = []
        for student in students_qs.select_related('classe').order_by('classe__name', 'last_name', 'first_name'):
            records = attendance_qs.filter(student=student)
            agg = records.aggregate(
                total=Count('id'),
                present=Count('id', filter=Q(status='PRESENT')),
                absent=Count('id', filter=Q(status='ABSENT')),
                late=Count('id', filter=Q(status='LATE')),
            )
            total_hours_absent = sum(r.hours_absent for r in records if r.hours_absent)
            result.append({
                'student_name': student.full_name,
                'classe': student.classe.name if student.classe else '—',
                'total_days': agg['total'],
                'present': agg['present'],
                'absent': agg['absent'],
                'late': agg['late'],
                'hours_absent': round(total_hours_absent, 1),
            })

        global_stats = attendance_qs.aggregate(
            total=Count('id'),
            present=Count('id', filter=Q(status='PRESENT')),
            absent=Count('id', filter=Q(status='ABSENT')),
            late=Count('id', filter=Q(status='LATE')),
        )

        pdf_buffer = _build_attendance_report_pdf(
            result, global_stats, start_date, end_date, classe_name, period
        )
        response = HttpResponse(pdf_buffer.getvalue(), content_type='application/pdf')
        filename = f"rapport_absences_{period or 'custom'}_{start_date}_{end_date}.pdf"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    def _resolve_date_range(self, period, date_from, date_to):
        from django.utils import timezone
        import calendar
        from datetime import timedelta, date

        if date_from or date_to:
            start = date.fromisoformat(date_from) if date_from else date.today()
            end = date.fromisoformat(date_to) if date_to else date.today()
            return start, end

        today = timezone.localdate()
        if period == 'daily':
            return today, today
        if period == 'weekly':
            start = today - timedelta(days=today.weekday())
            end = start + timedelta(days=6)
            return start, end
        if period == 'monthly':
            start = today.replace(day=1)
            _, last_day = calendar.monthrange(today.year, today.month)
            end = today.replace(day=last_day)
            return start, end
        return today, today


def _build_attendance_report_pdf(rows, global_stats, start_date, end_date, classe_name=None, period=None):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

    title = 'Rapport d\'absences'
    subtitle = f"Période : {start_date.strftime('%d/%m/%Y')} — {end_date.strftime('%d/%m/%Y')}"
    if period == 'daily':
        subtitle = f"Période : journalier — {start_date.strftime('%d/%m/%Y')}"
    elif period == 'weekly':
        subtitle = f"Période : hebdomadaire — {start_date.strftime('%d/%m/%Y')} à {end_date.strftime('%d/%m/%Y')}"
    elif period == 'monthly':
        subtitle = f"Période : mensuel — {start_date.strftime('%d/%m/%Y')} à {end_date.strftime('%d/%m/%Y')}"

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4,
                            rightMargin=1.5 * cm, leftMargin=1.5 * cm,
                            topMargin=1.5 * cm, bottomMargin=1.5 * cm)
    styles = getSampleStyleSheet()
    header_style = ParagraphStyle('Header', parent=styles['Heading1'], fontSize=18, spaceAfter=12)
    normal_style = styles['Normal']
    normal_style.spaceAfter = 8

    elements = [Paragraph(title, header_style), Paragraph(subtitle, normal_style)]
    if classe_name:
        elements.append(Paragraph(f"Classe : {classe_name}", normal_style))
    elements.append(Spacer(1, 10))

    summary_text = (
        f"Total enregistrements : {global_stats.get('total', 0)} · "
        f"Présents : {global_stats.get('present', 0)} · "
        f"Retards : {global_stats.get('late', 0)} · "
        f"Absents : {global_stats.get('absent', 0)}"
    )
    elements.append(Paragraph(summary_text, normal_style))
    elements.append(Spacer(1, 10))

    table_data = [[
        'Élève', 'Classe', 'Jours', 'Présents', 'Retards', 'Absents', 'Heures abs.'
    ]]
    for row in rows:
        table_data.append([
            row['student_name'], row['classe'], row['total_days'],
            row['present'], row['late'], row['absent'], row['hours_absent'],
        ])

    table = Table(table_data, repeatRows=1, hAlign='LEFT')
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e40af')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('ALIGN', (2, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cbd5e0')),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
    ]))
    elements.append(table)

    doc.build(elements)
    buffer.seek(0)
    return buffer


class PaymentReportView(APIView):
    """Payment rates per class."""
    permission_classes = [IsAdmin]

    def get(self, request):
        classe_id = request.query_params.get('classe_id')
        _eid = _ecole_id(request)

        classes_qs = Classe.objects.all()
        if _eid:
            classes_qs = classes_qs.filter(ecole_id=_eid)
        if classe_id:
            classes_qs = classes_qs.filter(id=classe_id)

        result = []
        for cls in classes_qs.prefetch_related('students'):
            students = cls.students.filter(is_active=True)
            total = students.count()
            paid = students.filter(payment_status='PAID').count()
            pending = students.filter(payment_status='PENDING').count()
            overdue = students.filter(payment_status='OVERDUE').count()

            payments = Payment.objects.filter(student__classe=cls)
            collected = payments.filter(status='PAID').aggregate(t=Sum('amount'))['t'] or 0
            pending_amt = payments.filter(status='PENDING').aggregate(t=Sum('amount'))['t'] or 0
            overdue_amt = payments.filter(status='OVERDUE').aggregate(t=Sum('amount'))['t'] or 0

            result.append({
                'classe_id': cls.id,
                'classe_name': cls.name,
                'total_students': total,
                'paid_count': paid,
                'pending_count': pending,
                'overdue_count': overdue,
                'rate': round(paid / total * 100, 1) if total > 0 else 0,
                'total_collected': float(collected),
                'total_pending': float(pending_amt),
                'total_overdue': float(overdue_amt),
            })

        return Response(result)


BACKUP_APPS = [
    'users', 'students', 'teachers', 'classes', 'subjects',
    'grades', 'attendance', 'payments', 'reports', 'payroll',
]


class BackupView(APIView):
    """Sauvegarde complète des données (JSON) — réservé à l'administrateur."""
    permission_classes = [IsPlatformAdmin]

    def get(self, request):
        from io import StringIO
        from django.core import management
        from datetime import date as _d
        buf = StringIO()
        management.call_command(
            'dumpdata', *BACKUP_APPS,
            '--exclude', 'reports.documenttemplate',  # gabarits volumineux optionnels
            indent=2, stdout=buf,
        )
        data = buf.getvalue().encode('utf-8')
        resp = HttpResponse(data, content_type='application/json')
        resp['Content-Disposition'] = f'attachment; filename="sauvegarde_schoolpro_{_d.today()}.json"'
        return resp


class RestoreView(APIView):
    """Restauration depuis un fichier de sauvegarde JSON — administrateur only.

    Opération sensible : remplace les données par celles du fichier. Exige
    `confirm=true` pour être exécutée."""
    permission_classes = [IsPlatformAdmin]
    parser_classes = [MultiPartParser, FormParser]

    def post(self, request):
        import os
        import tempfile
        from django.core import management

        if str(request.data.get('confirm', '')).lower() not in ('true', '1', 'yes'):
            return Response({'error': 'Confirmation requise (confirm=true).'}, status=400)
        f = request.FILES.get('file')
        if not f:
            return Response({'error': 'Fichier de sauvegarde requis.'}, status=400)
        if not f.name.lower().endswith('.json'):
            return Response({'error': 'Le fichier doit être au format .json.'}, status=400)

        tmp = tempfile.NamedTemporaryFile(suffix='.json', delete=False)
        try:
            for chunk in f.chunks():
                tmp.write(chunk)
            tmp.close()
            management.call_command('loaddata', tmp.name)
        except Exception as exc:
            return Response({'error': f'Échec de la restauration : {exc}'}, status=400)
        finally:
            try:
                os.unlink(tmp.name)
            except OSError:
                pass

        from apps.users.models import log_activity
        log_activity(request.user, 'LOGIN', 'Restauration de sauvegarde effectuée', request)
        return Response({'message': 'Restauration effectuée avec succès.'})


class ClassRosterPDFView(APIView):
    """Liste des élèves d'une classe (nom, prénom, date et lieu de naissance).
    Accessible aux professeurs, éducateurs et à l'administration."""
    permission_classes = [CanViewReports]

    def get(self, request, classe_id):
        eid = _ecole_id(request)
        classes = Classe.objects.select_related('academic_year').filter(id=classe_id)
        if eid:
            classes = classes.filter(ecole_id=eid)
        classe = classes.first()
        if not classe:
            return Response({'error': 'Classe non trouvée.'}, status=404)
        if not _teacher_can_access_classe(request.user, classe.id):
            return Response({'error': "Accès limité à vos classes assignées."}, status=403)
        students = list(
            Student.objects.filter(classe=classe, is_active=True).order_by('last_name', 'first_name')
        )
        pdf = generate_class_roster_pdf(classe, students)
        safe = classe.name.replace(' ', '_').replace('/', '-')
        response = HttpResponse(pdf, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="liste_eleves_{safe}.pdf"'
        return response


class TimetablePDFView(APIView):
    """Emploi du temps d'une classe en PDF.
    Accessible aux professeurs, éducateurs et à l'administration."""
    permission_classes = [CanViewReports]

    def get(self, request, classe_id):
        from apps.classes.models import ScheduleEntry
        eid = _ecole_id(request)
        classes = Classe.objects.select_related('academic_year').filter(id=classe_id)
        if eid:
            classes = classes.filter(ecole_id=eid)
        classe = classes.first()
        if not classe:
            return Response({'error': 'Classe non trouvée.'}, status=404)
        if not _teacher_can_access_classe(request.user, classe.id):
            return Response({'error': "Accès limité à vos classes assignées."}, status=403)
        entries = list(
            ScheduleEntry.objects.select_related('subject')
            .filter(classe=classe).order_by('day', 'start_time')
        )
        pdf = generate_timetable_pdf(classe, entries)
        safe = classe.name.replace(' ', '_').replace('/', '-')
        response = HttpResponse(pdf, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="emploi_du_temps_{safe}.pdf"'
        return response


class TeacherTimetablePDFView(APIView):
    """Emploi du temps d'un professeur. Un professeur obtient le sien ;
    admin/directeur/éducateur peuvent demander celui d'un professeur donné."""
    permission_classes = [CanViewReports]

    def get(self, request, teacher_id=None):
        from apps.teachers.models import Teacher
        from apps.classes.models import ScheduleEntry
        eid = _ecole_id(request)

        if teacher_id and request.user.role in ('ADMIN', 'DIRECTOR', 'EDUCATEUR'):
            qs = Teacher.objects.select_related('user').filter(id=teacher_id)
            if eid:
                qs = qs.filter(ecole_id=eid)
            teacher = qs.first()
        else:
            teacher = getattr(request.user, 'teacher_profile', None)

        if not teacher:
            return Response({'error': 'Professeur non trouvé.'}, status=404)

        entries = list(
            ScheduleEntry.objects.select_related('subject', 'classe')
            .filter(subject__teacher=teacher).order_by('day', 'start_time')
        )
        pdf = generate_teacher_timetable_pdf(teacher, entries)
        safe = (teacher.full_name or 'prof').replace(' ', '_').replace('/', '-')
        response = HttpResponse(pdf, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="emploi_du_temps_{safe}.pdf"'
        return response


class StudentSheetPDFView(APIView):
    """Fiche élève PDF (identité, scolarité, parent, situation financière)."""
    permission_classes = [CanViewReports]

    def get(self, request, student_id):
        eid = _ecole_id(request)
        students = Student.objects.select_related('classe', 'classe__academic_year').filter(id=student_id)
        if eid:
            students = students.filter(ecole_id=eid)
        student = students.first()
        if not student:
            return Response({'error': 'Élève non trouvé.'}, status=404)
        pdf = generate_student_sheet_pdf(student)
        response = HttpResponse(pdf, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="fiche_{student.matricule}.pdf"'
        return response


class ExcelExportView(APIView):
    """Export grades, absences, payments, rankings ou students to Excel."""
    permission_classes = [IsAdmin]

    def get(self, request):
        export_type = request.query_params.get('type', 'grades')
        classe_id = request.query_params.get('classe_id')
        _eid = _ecole_id(request)

        wb = openpyxl.Workbook()
        ws = wb.active

        # Styles
        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='1E40AF', end_color='1E40AF', fill_type='solid')
        header_align = Alignment(horizontal='center', vertical='center')
        thin = Side(style='thin', color='D1D5DB')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        def style_header(row_num, num_cols):
            for col in range(1, num_cols + 1):
                cell = ws.cell(row=row_num, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
                cell.border = border

        def style_row(row_num, num_cols, even):
            fill_color = 'F8FAFF' if even else 'FFFFFF'
            for col in range(1, num_cols + 1):
                cell = ws.cell(row=row_num, column=col)
                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
                cell.border = border
                cell.alignment = Alignment(vertical='center')

        if export_type == 'grades':
            ws.title = 'Notes'
            headers = ['Élève', 'Classe', 'Matière', 'Type', 'Note', 'Note Max', 'Note /20', 'Date']
            ws.append(headers)
            style_header(1, len(headers))

            qs = Grade.objects.select_related('student', 'student__classe', 'subject').all()
            if _eid:
                qs = qs.filter(student__ecole_id=_eid)
            if classe_id:
                qs = qs.filter(student__classe_id=classe_id)

            for i, g in enumerate(qs):
                row = [
                    g.student.full_name,
                    g.student.classe.name if g.student.classe else '—',
                    g.subject.name,
                    g.type_evaluation,
                    float(g.value),
                    float(g.max_value),
                    g.normalized_value,
                    str(g.date),
                ]
                ws.append(row)
                style_row(i + 2, len(headers), i % 2 == 0)

            col_widths = [25, 12, 20, 15, 8, 8, 8, 12]

        elif export_type == 'absences':
            ws.title = 'Absences'
            headers = ['Élève', 'Classe', 'Date', 'Heure Entrée', 'Heure Sortie', 'Statut', 'Heures Absence']
            ws.append(headers)
            style_header(1, len(headers))

            qs = Attendance.objects.select_related('student', 'student__classe').all()
            if _eid:
                qs = qs.filter(student__ecole_id=_eid)
            if classe_id:
                qs = qs.filter(student__classe_id=classe_id)

            for i, a in enumerate(qs):
                row = [
                    a.student.full_name,
                    a.student.classe.name if a.student.classe else '—',
                    str(a.date),
                    str(a.check_in) if a.check_in else '—',
                    str(a.check_out) if a.check_out else '—',
                    a.get_status_display(),
                    a.hours_absent,
                ]
                ws.append(row)
                style_row(i + 2, len(headers), i % 2 == 0)

            col_widths = [25, 12, 12, 12, 12, 12, 15]

        elif export_type == 'payments':
            ws.title = 'Paiements'
            headers = ['Élève', 'Classe', 'Type', 'Montant (FCFA)', 'Statut', 'Date Paiement', 'Échéance', 'N° Reçu']
            ws.append(headers)
            style_header(1, len(headers))

            qs = Payment.objects.select_related('student', 'student__classe').all()
            if _eid:
                qs = qs.filter(student__ecole_id=_eid)
            if classe_id:
                qs = qs.filter(student__classe_id=classe_id)

            for i, p in enumerate(qs):
                row = [
                    p.student.full_name,
                    p.student.classe.name if p.student.classe else '—',
                    p.get_payment_type_display(),
                    float(p.amount),
                    p.get_status_display(),
                    str(p.payment_date),
                    str(p.due_date) if p.due_date else '—',
                    p.receipt_number,
                ]
                ws.append(row)
                style_row(i + 2, len(headers), i % 2 == 0)

            col_widths = [25, 12, 20, 15, 12, 14, 12, 18]

        elif export_type == 'rankings':
            ws.title = 'Classement'
            headers = ['Rang', 'Élève', 'Classe', 'Moyenne Générale', 'Appréciation']
            ws.append(headers)
            style_header(1, len(headers))

            if not classe_id:
                return Response({'error': 'classe_id requis pour le classement'}, status=400)

            classe_qs = Classe.objects.filter(id=classe_id)
            if _eid:
                classe_qs = classe_qs.filter(ecole_id=_eid)
            if not classe_qs.exists():
                return Response({'error': 'Classe non trouvée.'}, status=404)
            students = Student.objects.filter(classe_id=classe_id, is_active=True)
            subjects = Subject.objects.filter(classe_id=classe_id)
            rankings = []
            for s in students:
                tw, tc = 0, 0
                for subj in subjects:
                    grades = Grade.objects.filter(student=s, subject=subj)
                    if grades.exists():
                        avg = grades.aggregate(a=DAvg('value'))['a']
                        tw += float(avg) * float(subj.coefficient)
                        tc += float(subj.coefficient)
                general_avg = round(tw / tc, 2) if tc > 0 else 0
                appreciation = (
                    'Très Bien' if general_avg >= 16 else
                    'Bien' if general_avg >= 14 else
                    'Assez Bien' if general_avg >= 12 else
                    'Passable' if general_avg >= 10 else 'Insuffisant'
                )
                rankings.append((s, general_avg, appreciation))

            rankings.sort(key=lambda x: x[1], reverse=True)
            for i, (s, avg, appr) in enumerate(rankings):
                row = [i + 1, s.full_name, s.classe.name if s.classe else '—', avg, appr]
                ws.append(row)
                style_row(i + 2, len(headers), i % 2 == 0)

            col_widths = [8, 28, 14, 18, 15]

        elif export_type == 'students':
            ws.title = 'Élèves'
            headers = ['Matricule', 'Nom', 'Prénom', 'Genre', 'Date de naissance',
                       'Lieu de naissance', 'Classe', 'Parent', 'Téléphone parent',
                       'Email parent', 'Statut paiement', 'Actif']
            ws.append(headers)
            style_header(1, len(headers))

            qs = Student.objects.select_related('classe').all()
            if _eid:
                qs = qs.filter(ecole_id=_eid)
            if classe_id:
                qs = qs.filter(classe_id=classe_id)
            qs = qs.order_by('last_name', 'first_name')

            pay_map = dict(Student.PaymentStatus.choices)
            for i, s in enumerate(qs):
                row = [
                    s.matricule, s.last_name, s.first_name, s.get_gender_display(),
                    str(s.date_of_birth) if s.date_of_birth else '—',
                    s.birth_place or '—',
                    s.classe.name if s.classe else '—',
                    s.parent_name or '—', s.parent_phone or '—', s.parent_email or '—',
                    pay_map.get(s.payment_status, s.payment_status),
                    'Oui' if s.is_active else 'Non',
                ]
                ws.append(row)
                style_row(i + 2, len(headers), i % 2 == 0)

            col_widths = [16, 16, 16, 10, 15, 16, 12, 20, 15, 22, 14, 8]

        else:
            return Response({'error': 'type invalide'}, status=400)

        # Set column widths
        for i, width in enumerate(col_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

        ws.row_dimensions[1].height = 22

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"export_{export_type}.xlsx"
        response = HttpResponse(
            output,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


class ClassStatsView(APIView):
    """Per-class statistics: student count, grade average, attendance rates."""
    permission_classes = [IsAdmin]

    def get(self, request):
        from django.utils import timezone
        today = timezone.now().date()

        classes = Classe.objects.select_related('level').all()
        _eid = _ecole_id(request)
        if _eid:
            classes = classes.filter(ecole_id=_eid)
        class_ids = list(classes.values_list('id', flat=True))

        # --- Bulk grade averages per class ---
        grade_rows = (
            Grade.objects
            .filter(student__classe_id__in=class_ids)
            .values('student__classe_id')
            .annotate(avg=Avg('value'))
        )
        grade_avg_map = {row['student__classe_id']: row['avg'] for row in grade_rows}

        # --- Bulk attendance totals per class (all time) ---
        attendance_rows = (
            Attendance.objects
            .filter(student__classe_id__in=class_ids)
            .values('student__classe_id')
            .annotate(
                total=Count('id'),
                absent=Count('id', filter=Q(status='ABSENT')),
            )
        )
        attendance_map = {
            row['student__classe_id']: row
            for row in attendance_rows
        }

        # --- Bulk today's attendance per class ---
        today_rows = (
            Attendance.objects
            .filter(student__classe_id__in=class_ids, date=today)
            .values('student__classe_id')
            .annotate(
                present_count=Count('id', filter=Q(status='PRESENT')),
                absent_count=Count('id', filter=Q(status='ABSENT')),
                late_count=Count('id', filter=Q(status='LATE')),
            )
        )
        today_map = {row['student__classe_id']: row for row in today_rows}

        # --- Bulk student counts per class ---
        student_counts = (
            Student.objects
            .filter(classe_id__in=class_ids, is_active=True)
            .values('classe_id')
            .annotate(count=Count('id'))
        )
        student_count_map = {row['classe_id']: row['count'] for row in student_counts}

        result = []
        for cls in classes:
            cid = cls.id

            raw_avg = grade_avg_map.get(cid)
            average = round(raw_avg, 1) if raw_avg is not None else None

            att = attendance_map.get(cid, {})
            total_att = att.get('total', 0) or 0
            absent_att = att.get('absent', 0) or 0
            absent_rate = round(absent_att / total_att * 100, 1) if total_att > 0 else 0

            td = today_map.get(cid, {})

            result.append({
                'id': cid,
                'name': cls.name,
                'level_id': cls.level_id,
                'level_name': cls.level.name if cls.level else None,
                'student_count': student_count_map.get(cid, 0),
                'average': average,
                'absent_rate': absent_rate,
                'present_count': td.get('present_count', 0) or 0,
                'absent_count': td.get('absent_count', 0) or 0,
                'late_count': td.get('late_count', 0) or 0,
            })

        return Response(result)


def _platform_settings(request=None):
    """Return the matching PlatformSettings object or global fallback."""
    try:
        from .models import PlatformSettings
        ecole = _selected_ecole(request) if request is not None else None
        if ecole:
            obj = PlatformSettings.objects.filter(ecole=ecole).first()
            if obj:
                return obj
        return PlatformSettings.get_solo(ecole=None)
    except Exception:
        return None


def _get_platform_school_type(request=None):
    """Return the school_type stored in platform settings (DB), defaulting to 'all'."""
    try:
        ps = _platform_settings(request)
        if ps and ps.school_type:
            return ps.school_type
    except Exception:
        pass
    try:
        from django.conf import settings as _s
        return getattr(_s, 'SCHOOL_TYPE', 'all') or 'all'
    except Exception:
        return 'all'


class PlatformSettingsView(APIView):
    """Réglages d'identité de l'établissement.
    Lecture : tout utilisateur authentifié. Écriture : admin/directeur."""
    parser_classes = [MultiPartParser, FormParser, JSONParser]

    def get_permissions(self):
        if self.request.method in ('GET', 'HEAD', 'OPTIONS'):
            return [IsAuthenticated()]
        return [IsAdmin()]

    def get(self, request):
        from .serializers import PlatformSettingsSerializer
        obj = _platform_settings(request)
        if obj is None:
            obj = _platform_settings(None)
        return Response(PlatformSettingsSerializer(obj, context={'request': request}).data)

    def put(self, request):
        return self._update(request)

    def patch(self, request):
        return self._update(request)

    def _update(self, request):
        from .serializers import PlatformSettingsSerializer
        ecole = _selected_ecole(request)
        obj = PlatformSettings.get_solo(ecole=ecole)
        ser = PlatformSettingsSerializer(obj, data=request.data, partial=True, context={'request': request})
        ser.is_valid(raise_exception=True)
        ser.save()
        return Response(ser.data)


def _resolve_template(document_type, template_id=None, school_type=None, request=None):
    """Return template config dict or None. Prefers school-specific and school-type defaults."""
    from .models import DocumentTemplate
    selected_ecole = _selected_ecole(request) if request is not None else None
    allowed_filter = Q(ecole__isnull=True)
    if selected_ecole is not None:
        allowed_filter = Q(ecole=selected_ecole) | Q(ecole__isnull=True)

    # template_id peut arriver sous forme de chaîne ('null', '') — on sécurise
    if template_id not in (None, '', 'null', 'undefined'):
        try:
            tmpl = DocumentTemplate.objects.filter(
                allowed_filter,
                id=int(template_id),
                document_type=document_type,
            ).first()
        except (ValueError, TypeError):
            tmpl = None
        if tmpl:
            return tmpl.config

    st = school_type or _get_platform_school_type(request)
    tmpl = (
        DocumentTemplate.objects.filter(allowed_filter, document_type=document_type, school_type=st, is_default=True)
        .order_by('-ecole_id').first()
        or DocumentTemplate.objects.filter(allowed_filter, document_type=document_type, school_type='all', is_default=True)
        .order_by('-ecole_id').first()
        or DocumentTemplate.objects.filter(allowed_filter, document_type=document_type, is_default=True)
        .order_by('-ecole_id').first()
    )
    return tmpl.config if tmpl else None


# ────────────────────────────────────────────────────────────────
# Document Template CRUD
# ────────────────────────────────────────────────────────────────

class DocumentTemplateListCreateView(APIView):
    permission_classes = [IsAdmin]

    def get(self, request):
        from .models import DocumentTemplate
        from .serializers import DocumentTemplateSerializer
        doc_type = request.query_params.get('document_type')
        school_type = request.query_params.get('school_type')
        selected_ecole = _selected_ecole(request)
        qs = DocumentTemplate.objects.filter(
            Q(ecole__isnull=True) if selected_ecole is None else (Q(ecole=selected_ecole) | Q(ecole__isnull=True))
        )
        if request.user.role != 'ADMIN' and getattr(request.user, 'ecole_id', None):
            qs = qs.filter(Q(ecole_id=request.user.ecole_id) | Q(ecole__isnull=True))
        if doc_type:
            qs = qs.filter(document_type=doc_type)
        if school_type:
            qs = qs.filter(school_type=school_type)
        qs = qs.order_by('-ecole_id', '-created_at')
        return Response(DocumentTemplateSerializer(qs, many=True).data)

    def post(self, request):
        from .models import DocumentTemplate
        from .serializers import DocumentTemplateSerializer
        selected_ecole = _selected_ecole(request)
        if request.user.role != 'ADMIN' and getattr(request.user, 'ecole', None):
            selected_ecole = request.user.ecole
        ser = DocumentTemplateSerializer(data=request.data)
        if ser.is_valid():
            ser.save(created_by=request.user, ecole=selected_ecole)
            return Response(ser.data, status=201)
        return Response(ser.errors, status=400)


class DocumentTemplateDetailView(APIView):
    permission_classes = [IsAdmin]

    def _get_obj(self, pk, request):
        from .models import DocumentTemplate
        selected_ecole = _selected_ecole(request)
        qs = DocumentTemplate.objects.filter(pk=pk)
        if selected_ecole is not None:
            qs = qs.filter(Q(ecole=selected_ecole) | Q(ecole__isnull=True))
        elif request.user.role != 'ADMIN' and getattr(request.user, 'ecole_id', None):
            qs = qs.filter(Q(ecole_id=request.user.ecole_id) | Q(ecole__isnull=True))
        try:
            return qs.first()
        except DocumentTemplate.DoesNotExist:
            return None

    def get(self, request, pk):
        from .serializers import DocumentTemplateSerializer
        obj = self._get_obj(pk, request)
        if not obj:
            return Response({'error': 'Modèle introuvable.'}, status=404)
        return Response(DocumentTemplateSerializer(obj).data)

    def put(self, request, pk):
        from .serializers import DocumentTemplateSerializer
        obj = self._get_obj(pk, request)
        if not obj:
            return Response({'error': 'Modèle introuvable.'}, status=404)
        ser = DocumentTemplateSerializer(obj, data=request.data, partial=True)
        if ser.is_valid():
            ser.save()
            return Response(ser.data)
        return Response(ser.errors, status=400)

    def delete(self, request, pk):
        obj = self._get_obj(pk, request)
        if not obj:
            return Response({'error': 'Modèle introuvable.'}, status=404)
        obj.delete()
        return Response(status=204)


# ────────────────────────────────────────────────────────────────
# Receipt & Card PDF with template support
# ────────────────────────────────────────────────────────────────

class ClassListPDFView(APIView):
    """Liste d'une classe en PDF, par ordre de mérite ou alphabétique.
    Accessible aux professeurs, éducateurs et à l'administration."""
    permission_classes = [CanViewReports]

    def get(self, request, classe_id):
        order = request.query_params.get('order', 'merit')
        if order not in ('merit', 'alpha'):
            order = 'merit'
        eid = _ecole_id(request)
        classes = Classe.objects.select_related('level', 'academic_year').filter(id=classe_id)
        if eid:
            classes = classes.filter(ecole_id=eid)
        classe = classes.first()
        if not classe:
            return Response({'error': 'Classe non trouvée.'}, status=404)
        if not _teacher_can_access_classe(request.user, classe.id):
            return Response({'error': "Accès limité à vos classes assignées."}, status=403)

        students = Student.objects.filter(classe=classe, is_active=True)
        subjects = list(Subject.objects.filter(classe=classe).order_by('name'))
        subj_meta = [{'name': s.name, 'coefficient': float(s.coefficient)} for s in subjects]

        rows = []
        for s in students:
            averages = []
            total_weighted, total_coeff = 0, 0
            for subj in subjects:
                grades = Grade.objects.filter(student=s, subject=subj)
                if grades.exists():
                    a = round(float(grades.aggregate(a=Avg('value'))['a']), 2)
                    averages.append(a)
                    total_weighted += a * float(subj.coefficient)
                    total_coeff += float(subj.coefficient)
                else:
                    averages.append(None)
            general = round(total_weighted / total_coeff, 2) if total_coeff > 0 else None
            rows.append({
                'name': s.full_name,
                'averages': averages,
                'total': round(total_weighted, 2),
                'average': general,
            })

        if order == 'alpha':
            rows.sort(key=lambda r: r['name'].lower())
        else:
            rows.sort(key=lambda r: (r['average'] is None, -(r['average'] or 0)))

        pdf = generate_class_list_pdf(classe, subj_meta, rows, order)
        safe = classe.name.replace(' ', '_').replace('/', '-')
        response = HttpResponse(pdf, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="liste_{safe}_{order}.pdf"'
        return response


class SubjectSheetPDFView(APIView):
    """Feuille de notes (Nom & Prénom · Notes saisies · Moyenne) d'une matière
    pour une classe — imprimable par le professeur."""
    permission_classes = [IsAdminOrTeacher]

    def get(self, request):
        subject_id = request.query_params.get('subject_id')
        classe_id = request.query_params.get('classe_id')
        order = request.query_params.get('order', 'merit')
        if order not in ('merit', 'alpha'):
            order = 'merit'
        if not subject_id or not classe_id:
            return Response({'error': 'subject_id et classe_id requis.'}, status=400)
        eid = _ecole_id(request)
        try:
            subject = Subject.objects.get(id=subject_id)
            classes = Classe.objects.select_related('academic_year').filter(id=classe_id)
            if eid:
                classes = classes.filter(ecole_id=eid)
            classe = classes.first()
            if not classe:
                raise Classe.DoesNotExist
        except Subject.DoesNotExist:
            return Response({'error': 'Matière introuvable.'}, status=404)
        except Classe.DoesNotExist:
            return Response({'error': 'Classe introuvable.'}, status=404)

        students = Student.objects.filter(classe=classe, is_active=True).order_by('last_name', 'first_name')
        rows = []
        for s in students:
            grades = Grade.objects.filter(student=s, subject=subject).order_by('date')
            gl = [{'value': float(g.value), 'max': float(g.max_value), 'type': g.type_evaluation}
                  for g in grades]
            if gl:
                avg = round(sum(g['value'] * 20 / g['max'] if g['max'] else 0 for g in gl) / len(gl), 2)
                total_value = round(sum(g['value'] for g in gl), 2)
                total_max = round(sum(g['max'] for g in gl), 2)
            else:
                avg = None
                total_value = total_max = None
            rows.append({
                'name': s.full_name, 'grades': gl, 'average': avg,
                'total_value': total_value, 'total_max': total_max,
            })

        # Rang au mérite (sur la moyenne de la matière) — conservé même en tri alphabétique
        ranked = sorted(rows, key=lambda r: (r['average'] is None, -(r['average'] or 0)))
        for i, r in enumerate(ranked, start=1):
            r['rank'] = i if r['average'] is not None else None
        if order == 'alpha':
            rows.sort(key=lambda r: r['name'].lower())
        else:
            rows = ranked

        pdf = generate_subject_sheet_pdf(classe, subject, rows, order)
        safe = f"{subject.name}_{classe.name}_{order}".replace(' ', '_').replace('/', '-')
        response = HttpResponse(pdf, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="notes_{safe}.pdf"'
        return response


class ReceiptPDFView(APIView):
    permission_classes = [IsCashManager]

    def get(self, request, payment_id):
        eid = _ecole_id(request)
        try:
            payment = Payment.objects.select_related(
                'student', 'student__classe', 'recorded_by'
            ).get(id=payment_id)
        except Payment.DoesNotExist:
            return Response({'error': 'Paiement non trouvé.'}, status=404)
        if eid and not (
            payment.ecole_id == eid or
            (payment.student and payment.student.ecole_id == eid)
        ):
            return Response({'error': 'Paiement non trouvé.'}, status=404)

        template_config = _resolve_template(
            'receipt',
            template_id=request.query_params.get('template_id'),
            school_type=request.query_params.get('school_type'),
            request=request,
        ) or {}

        pdf_buffer = generate_receipt_pdf_templated(payment, template_config)
        response = HttpResponse(pdf_buffer, content_type='application/pdf')
        response['Content-Disposition'] = (
            f'attachment; filename="recu_{payment.receipt_number}.pdf"'
        )
        return response


class StudentCardPDFView(APIView):
    permission_classes = [IsAdminOrTeacher]

    def get(self, request, student_id):
        eid = _ecole_id(request)
        students = Student.objects.select_related(
            'classe', 'classe__academic_year'
        ).filter(id=student_id)
        if eid:
            students = students.filter(ecole_id=eid)
        student = students.first()
        if not student:
            return Response({'error': 'Élève non trouvé.'}, status=404)

        template_config = _resolve_template(
            'card',
            template_id=request.query_params.get('template_id'),
            school_type=request.query_params.get('school_type'),
            request=request,
        ) or {}

        pdf_buffer = generate_card_pdf_templated(student, template_config)
        response = HttpResponse(pdf_buffer, content_type='application/pdf')
        response['Content-Disposition'] = (
            f'attachment; filename="carte_{student.matricule}.pdf"'
        )
        return response
