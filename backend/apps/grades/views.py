from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.exceptions import PermissionDenied
from django.db.models import Avg, Count
from apps.users.permissions import CanEditGrades
from apps.users.mixins import EcoleScopedMixin
from .models import Grade
from .serializers import GradeSerializer, GradeCreateSerializer, BulkGradeSerializer
import logging

logger = logging.getLogger('apps')


class GradeViewSet(EcoleScopedMixin, viewsets.ModelViewSet):
    ecole_lookup = 'student__ecole'
    queryset = Grade.objects.select_related(
        'student', 'subject', 'subject__classe', 'created_by'
    ).all()
    permission_classes = [CanEditGrades]
    filterset_fields = ['student', 'subject', 'type_evaluation', 'subject__classe', 'period']
    search_fields = ['student__first_name', 'student__last_name']
    ordering_fields = ['date', 'value']

    def _check_teacher_owns_subject(self, subject_id):
        """Ensure a TEACHER can only write grades for their own subjects."""
        if self.request.user.role != 'TEACHER':
            return
        try:
            teacher = self.request.user.teacher_profile
        except Exception:
            raise PermissionDenied("Profil professeur introuvable.")
        if not teacher.subjects.filter(id=subject_id).exists():
            raise PermissionDenied("Vous ne pouvez saisir des notes que pour vos propres matières.")

    def get_serializer_class(self):
        if self.action in ['create']:
            return GradeCreateSerializer
        return GradeSerializer

    def _record_history(self, instance, change_type, old_value=None):
        from .models import GradeHistory
        GradeHistory.objects.create(
            grade=instance if change_type != 'DELETED' else None,
            student_name=instance.student.full_name,
            subject_name=instance.subject.name,
            changed_by=self.request.user,
            change_type=change_type,
            old_value=old_value,
            new_value=instance.value if change_type != 'DELETED' else old_value,
        )

    def _log(self, action, desc, old_value=None, new_value=None):
        from apps.users.models import log_activity
        log_activity(self.request.user, action, desc, self.request,
                     old_value=old_value, new_value=new_value)

    def perform_create(self, serializer):
        subject_id = serializer.validated_data['subject'].id
        self._check_teacher_owns_subject(subject_id)
        serializer.save(created_by=self.request.user)
        g = serializer.instance
        self._record_history(g, 'CREATED')
        self._log('GRADE_ADD', f"Note {g.value}/{g.max_value} ({g.type_evaluation}) — {g.student} — {g.subject.name}")

    def perform_update(self, serializer):
        subject_id = serializer.instance.subject_id
        self._check_teacher_owns_subject(subject_id)
        old_value = serializer.instance.value
        serializer.save()
        g = serializer.instance
        self._record_history(g, 'UPDATED', old_value)
        self._log('GRADE_EDIT', f"Note modifiée — {g.student} — {g.subject.name}",
                  old_value=f"{old_value}/{g.max_value}", new_value=f"{g.value}/{g.max_value}")

    def perform_destroy(self, instance):
        self._check_teacher_owns_subject(instance.subject_id)
        self._record_history(instance, 'DELETED', instance.value)
        self._log('GRADE_DELETE', f"Note {instance.value} supprimée — {instance.student} — {instance.subject.name}")
        instance.delete()

    @action(detail=False, methods=['post'])
    def bulk_create(self, request):
        """Create multiple grades at once."""
        serializer = BulkGradeSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        grades = []
        for grade_data in serializer.validated_data['grades']:
            grade_data['created_by'] = request.user
            grades.append(Grade(**grade_data))
        Grade.objects.bulk_create(grades)
        return Response(
            {'message': f'{len(grades)} notes enregistrées avec succès.'},
            status=status.HTTP_201_CREATED
        )

    @action(detail=False, methods=['get'])
    def import_template(self, request):
        """Modèle Excel d'import de notes (matricule, matière, note, sur, période, type)."""
        import io
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from django.http import HttpResponse
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Notes'
        headers = ['matricule', 'matiere', 'note', 'sur', 'periode', 'type']
        for i, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=i, value=h.upper())
            c.font = Font(bold=True, color='FFFFFF')
            c.fill = PatternFill('solid', fgColor='1E40AF')
            c.alignment = Alignment(horizontal='center')
        ws.append(['ELV-XXXX', 'Mathématiques', 15, 20, 1, 'Devoir'])
        for i, w in enumerate([18, 22, 8, 8, 10, 16], 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        resp = HttpResponse(buf, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        resp['Content-Disposition'] = 'attachment; filename="modele_import_notes.xlsx"'
        return resp

    @action(detail=False, methods=['post'],
            parser_classes=[MultiPartParser, FormParser])
    def import_grades(self, request):
        """Import de notes depuis un fichier .xlsx/.csv.

        Colonnes : matricule, matiere, note, sur (déf. 20), periode (déf. 1),
        type (déf. Devoir). Les élèves sont retrouvés par matricule et les
        matières par nom dans la classe de l'élève."""
        from apps.students.views import _parse_file
        from apps.students.models import Student
        from apps.subjects.models import Subject

        file = request.FILES.get('file')
        if not file:
            return Response({'error': 'Aucun fichier fourni.'}, status=400)
        ext = file.name.rsplit('.', 1)[-1].lower()
        if ext not in ('xlsx', 'csv'):
            return Response({'error': 'Format non supporté (.xlsx ou .csv).'}, status=400)
        try:
            rows = _parse_file(file, ext)
        except Exception as exc:
            return Response({'error': f'Erreur de lecture: {exc}'}, status=400)

        eid = None
        u = request.user
        if u.role != 'ADMIN':
            eid = getattr(u, 'ecole_id', None)

        created, errors = 0, []
        for n, row in enumerate(rows, start=2):
            row = {k.strip().lower(): (str(v).strip() if v is not None else '') for k, v in row.items()}
            mat = row.get('matricule', '')
            subj_name = row.get('matiere') or row.get('matière', '')
            note = row.get('note', '')
            if not mat or not subj_name or note == '':
                errors.append({'row': n, 'message': 'matricule, matiere et note requis'})
                continue
            students = Student.objects.filter(matricule__iexact=mat)
            if eid:
                students = students.filter(ecole_id=eid)
            student = students.first()
            if not student:
                errors.append({'row': n, 'message': f'Élève introuvable: {mat}'})
                continue
            subject = Subject.objects.filter(classe=student.classe, name__iexact=subj_name).first()
            if not subject:
                errors.append({'row': n, 'message': f'Matière introuvable dans la classe: {subj_name}'})
                continue
            try:
                value = float(str(note).replace(',', '.'))
                max_value = float(str(row.get('sur') or 20).replace(',', '.'))
                period = int(row.get('periode') or 1)
            except ValueError:
                errors.append({'row': n, 'message': 'Valeurs numériques invalides'})
                continue
            if value < 0 or value > max_value:
                errors.append({'row': n, 'message': f'Note hors bornes (0..{max_value:g})'})
                continue
            self._check_teacher_owns_subject(subject.id)
            Grade.objects.create(
                student=student, subject=subject, value=value, max_value=max_value,
                period=period, type_evaluation=(row.get('type') or 'Devoir'),
                created_by=request.user,
            )
            created += 1

        return Response({'created': created, 'errors': errors, 'total': created + len(errors)})

    @action(detail=False, methods=['get'])
    def by_subject_class(self, request):
        """Get grades filtered by subject and class."""
        subject_id = request.query_params.get('subject_id')
        classe_id = request.query_params.get('classe_id')
        if not subject_id:
            return Response({'error': 'subject_id requis'}, status=400)
        grades = self.queryset.filter(subject_id=subject_id)
        if classe_id:
            grades = grades.filter(student__classe_id=classe_id)
        serializer = GradeSerializer(grades, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def averages(self, request):
        """Calculate averages per student per subject for a class."""
        classe_id = request.query_params.get('classe_id')
        subject_id = request.query_params.get('subject_id')
        if not classe_id:
            return Response({'error': 'classe_id requis'}, status=400)

        filters = {'student__classe_id': classe_id}
        if subject_id:
            filters['subject_id'] = subject_id

        averages = (
            Grade.objects.filter(**filters)
            .values('student__id', 'student__last_name', 'student__first_name', 'subject__name')
            .annotate(average=Avg('value'), grade_count=Count('id'))
            .order_by('student__last_name')
        )
        result = [
            {
                'student_id': avg['student__id'],
                'student_name': f"{avg['student__last_name']} {avg['student__first_name']}",
                'subject_name': avg['subject__name'],
                'average': round(avg['average'], 2),
                'grade_count': avg['grade_count'],
            }
            for avg in averages
        ]
        return Response(result)

    @action(detail=False, methods=['get'])
    def ranking(self, request):
        """Get student ranking for a class."""
        classe_id = request.query_params.get('classe_id')
        if not classe_id:
            return Response({'error': 'classe_id requis'}, status=400)

        from apps.subjects.models import Subject
        from apps.students.models import Student

        period = request.query_params.get('period')
        try:
            period = int(period) if period not in (None, '', 'annual') else None
        except (ValueError, TypeError):
            period = None

        students = Student.objects.filter(classe_id=classe_id, is_active=True)
        subjects = Subject.objects.filter(classe_id=classe_id)

        rankings = []
        for student in students:
            total_weighted = 0
            total_coeff = 0
            subject_averages = []
            for subject in subjects:
                grades = Grade.objects.filter(student=student, subject=subject)
                if period:
                    grades = grades.filter(period=period)
                if grades.exists():
                    avg = grades.aggregate(avg=Avg('value'))['avg']
                    coeff = float(subject.coefficient)
                    total_weighted += float(avg) * coeff
                    total_coeff += coeff
                    subject_averages.append({
                        'subject': subject.name,
                        'average': round(float(avg), 2),
                        'coefficient': coeff,
                    })

            general_avg = round(total_weighted / total_coeff, 2) if total_coeff > 0 else 0
            rankings.append({
                'student_id': student.id,
                'student_name': student.full_name,
                'general_average': general_avg,
                'subject_averages': subject_averages,
            })

        rankings.sort(key=lambda x: x['general_average'], reverse=True)
        for i, r in enumerate(rankings):
            r['rank'] = i + 1
            if r['general_average'] >= 16:
                r['appreciation'] = 'Très Bien'
            elif r['general_average'] >= 14:
                r['appreciation'] = 'Bien'
            elif r['general_average'] >= 12:
                r['appreciation'] = 'Assez Bien'
            elif r['general_average'] >= 10:
                r['appreciation'] = 'Passable'
            else:
                r['appreciation'] = 'Insuffisant'

        return Response(rankings)

    @action(detail=False, methods=['get'])
    def student_grades(self, request):
        """Get all grades for a specific student."""
        student_id = request.query_params.get('student_id')
        if not student_id:
            return Response({'error': 'student_id requis'}, status=400)
        grades = self.queryset.filter(student_id=student_id)
        return Response(GradeSerializer(grades, many=True).data)

    @action(detail=False, methods=['get'])
    def history(self, request):
        """Grade modification history — filtered by subject or student."""
        from .models import GradeHistory
        qs = GradeHistory.objects.select_related('changed_by', 'grade').all()
        subject_id = request.query_params.get('subject_id')
        student_id = request.query_params.get('student_id')
        if subject_id:
            qs = qs.filter(grade__subject_id=subject_id)
        if student_id:
            qs = qs.filter(grade__student_id=student_id)
        data = [{
            'id': h.id,
            'student_name': h.student_name,
            'subject_name': h.subject_name,
            'change_type': h.change_type,
            'change_label': h.get_change_type_display(),
            'old_value': float(h.old_value) if h.old_value is not None else None,
            'new_value': float(h.new_value) if h.new_value is not None else None,
            'changed_by': h.changed_by.get_full_name() if h.changed_by else '—',
            'changed_at': h.changed_at.strftime('%d/%m/%Y %H:%M'),
        } for h in qs[:100]]
        return Response(data)
