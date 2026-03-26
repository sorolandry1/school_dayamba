import csv
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser, FormParser
from apps.users.permissions import IsAdmin, IsAdminOrReadOnly
from utils.qr_generator import generate_qr_for_student
from .models import Student
from .serializers import StudentListSerializer, StudentDetailSerializer, StudentCreateSerializer

# ── Colonnes attendues dans le fichier d'import ───────────────────────────────
IMPORT_COLUMNS = [
    ('nom',           'last_name',    True),   # (header, field, required)
    ('prenom',        'first_name',   True),
    ('genre',         'gender',       True),   # M ou F
    ('classe',        'classe_name',  True),   # nom de la classe, ex: "6ème A"
    ('date_naissance','date_of_birth',False),  # YYYY-MM-DD
    ('lieu_naissance','birth_place',  False),
    ('nationalite',   'nationality',  False),
    ('nom_parent',    'parent_name',  False),
    ('tel_parent',    'parent_phone', False),
    ('email_parent',  'parent_email', False),
    ('adresse',       'address',      False),
]
HEADER_LABELS = [c[0] for c in IMPORT_COLUMNS]


class StudentViewSet(viewsets.ModelViewSet):
    queryset = Student.objects.select_related('classe', 'classe__level').all()
    permission_classes = [IsAdminOrReadOnly]
    filterset_fields = ['classe', 'payment_status', 'is_active', 'gender']
    search_fields = ['first_name', 'last_name', 'matricule', 'parent_phone']
    ordering_fields = ['last_name', 'enrolled_date']

    def get_serializer_class(self):
        if self.action == 'list':
            return StudentListSerializer
        if self.action == 'create':
            return StudentCreateSerializer
        return StudentDetailSerializer

    def perform_create(self, serializer):
        student = serializer.save()
        generate_qr_for_student(student)

    @action(detail=True, methods=['post'])
    def generate_qr(self, request, pk=None):
        student = self.get_object()
        generate_qr_for_student(student)
        return Response({
            'message': 'QR Code généré avec succès.',
            'qr_code_data': student.qr_code_data
        })

    @action(detail=False, methods=['get'])
    def by_class(self, request):
        classe_id = request.query_params.get('classe_id')
        if not classe_id:
            return Response({'error': 'classe_id requis'}, status=status.HTTP_400_BAD_REQUEST)
        students = self.queryset.filter(classe_id=classe_id)
        serializer = StudentListSerializer(students, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=['get'])
    def situation(self, request, pk=None):
        """Return complete student situation: profile, grades, attendance, payments."""
        student = self.get_object()

        # Grades per subject
        from apps.grades.models import Grade
        from apps.subjects.models import Subject
        from django.db.models import Avg
        subjects = Subject.objects.filter(classe=student.classe)
        subject_data = []
        total_weighted, total_coeff = 0, 0
        for subj in subjects:
            grades = Grade.objects.filter(student=student, subject=subj)
            avg = None
            if grades.exists():
                avg = round(float(grades.aggregate(a=Avg('value'))['a']), 2)
                total_weighted += avg * float(subj.coefficient)
                total_coeff += float(subj.coefficient)
            subject_data.append({
                'subject': subj.name,
                'coefficient': float(subj.coefficient),
                'average': avg,
                'grades': [{'value': g.value, 'max_value': g.max_value,
                            'type': g.type_evaluation, 'date': str(g.date)} for g in grades],
            })
        general_avg = round(total_weighted / total_coeff, 2) if total_coeff > 0 else None

        # Attendance summary
        from apps.attendance.models import Attendance
        from django.db.models import Count, Q
        att = Attendance.objects.filter(student=student)
        att_summary = att.aggregate(
            total=Count('id'),
            present=Count('id', filter=Q(status='PRESENT')),
            late=Count('id', filter=Q(status='LATE')),
            absent=Count('id', filter=Q(status='ABSENT')),
        )

        # Payments
        from apps.payments.models import Payment
        from django.db.models import Sum
        payments = Payment.objects.filter(student=student).order_by('-payment_date')
        total_paid = payments.filter(status='PAID').aggregate(t=Sum('amount'))['t'] or 0
        payment_data = [{
            'id': p.id,
            'receipt_number': p.receipt_number,
            'amount': float(p.amount),
            'payment_type': p.payment_type,
            'status': p.status,
            'payment_date': str(p.payment_date),
            'due_date': str(p.due_date) if p.due_date else None,
        } for p in payments]

        return Response({
            'student': StudentDetailSerializer(student, context={'request': request}).data,
            'academic': {
                'general_average': general_avg,
                'subjects': subject_data,
            },
            'attendance': att_summary,
            'payments': {
                'total_paid': float(total_paid),
                'history': payment_data,
            },
        })

    # ── Import template ────────────────────────────────────────────────────────

    @action(detail=False, methods=['get'], permission_classes=[IsAdmin])
    def import_template(self, request):
        """Download a styled .xlsx template ready to fill in."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Élèves'

        # ── Styles ────────────────────────────────────────────────────────────
        header_fill  = PatternFill('solid', fgColor='1a365d')
        required_fill = PatternFill('solid', fgColor='2b6cb0')
        header_font  = Font(color='FFFFFF', bold=True, size=11)
        center       = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin         = Side(border_style='thin', color='CBD5E0')
        border       = Border(left=thin, right=thin, top=thin, bottom=thin)

        REQUIRED_COLS = {c[0] for c in IMPORT_COLUMNS if c[2]}

        # ── Header row ────────────────────────────────────────────────────────
        for col_idx, (header, _, required) in enumerate(IMPORT_COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header.upper())
            cell.fill   = required_fill if required else header_fill
            cell.font   = header_font
            cell.alignment = center
            cell.border = border

        # ── Example row ───────────────────────────────────────────────────────
        example = [
            'KONE', 'Ibrahim', 'M', '6ème A',
            '2012-05-14', 'Ouagadougou', 'Burkinabè',
            'Kone Moussa', '70123456', 'parent@mail.com', 'Secteur 12',
        ]
        for col_idx, val in enumerate(example, start=1):
            cell = ws.cell(row=2, column=col_idx, value=val)
            cell.fill   = PatternFill('solid', fgColor='EBF8FF')
            cell.alignment = center
            cell.border = border

        # ── Legend row ────────────────────────────────────────────────────────
        ws.cell(row=3, column=1, value='← Colonnes en bleu foncé = facultatives')
        ws.cell(row=3, column=1).font = Font(italic=True, color='667085', size=9)
        ws.cell(row=3, column=4, value='← Nom exact de la classe (ex: 6ème A)')
        ws.cell(row=3, column=4).font = Font(italic=True, color='667085', size=9)

        # ── Dropdown validation: genre ────────────────────────────────────────
        from openpyxl.worksheet.datavalidation import DataValidation
        dv = DataValidation(type='list', formula1='"M,F"', allow_blank=False)
        dv.sqref = f'C2:C1000'
        ws.add_data_validation(dv)

        # ── Column widths ────────────────────────────────────────────────────
        widths = [16, 16, 8, 12, 16, 18, 14, 20, 14, 24, 20]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.row_dimensions[1].height = 28

        # ── Freeze header ────────────────────────────────────────────────────
        ws.freeze_panes = 'A2'

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        response = HttpResponse(
            buf,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="modele_import_eleves.xlsx"'
        return response

    # ── Bulk import ────────────────────────────────────────────────────────────

    @action(detail=False, methods=['post'], permission_classes=[IsAdmin],
            parser_classes=[MultiPartParser, FormParser])
    def import_students(self, request):
        """Import students from an Excel (.xlsx) or CSV file.

        Returns a JSON report:
            { created, skipped, errors: [{row, message}], total }
        """
        file = request.FILES.get('file')
        if not file:
            return Response({'error': 'Aucun fichier fourni.'}, status=400)

        ext = file.name.rsplit('.', 1)[-1].lower()
        if ext not in ('xlsx', 'csv'):
            return Response({'error': 'Format non supporté. Utilisez .xlsx ou .csv.'}, status=400)

        # ── Parse file into list of dicts ──────────────────────────────────
        try:
            rows = _parse_file(file, ext)
        except Exception as exc:
            return Response({'error': f'Erreur de lecture du fichier: {exc}'}, status=400)

        if not rows:
            return Response({'error': 'Le fichier est vide.'}, status=400)

        # ── Validate headers ───────────────────────────────────────────────
        file_headers = [h.strip().lower() for h in rows[0].keys()]
        missing = [h for h in HEADER_LABELS if h not in file_headers]
        if missing:
            return Response({
                'error': f'Colonnes manquantes: {", ".join(missing)}. '
                         f'Téléchargez le modèle pour connaître le format attendu.'
            }, status=400)

        # ── Build classe lookup ────────────────────────────────────────────
        from apps.classes.models import Classe
        classes_map = {c.name.strip().lower(): c for c in Classe.objects.all()}

        created_count = 0
        skipped_count = 0
        errors = []

        for row_num, row in enumerate(rows, start=2):   # row 1 = header
            # Normalize keys
            row = {k.strip().lower(): (v or '').strip() for k, v in row.items()}

            # Required field check
            missing_required = [
                col for col, _, req in IMPORT_COLUMNS
                if req and not row.get(col)
            ]
            if missing_required:
                errors.append({
                    'row': row_num,
                    'message': f'Champs obligatoires manquants: {", ".join(missing_required)}',
                    'data': dict(list(row.items())[:4]),
                })
                continue

            # Genre validation
            genre = row.get('genre', '').upper()
            if genre not in ('M', 'F'):
                errors.append({
                    'row': row_num,
                    'message': f'Genre invalide "{genre}". Utiliser M ou F.',
                    'data': dict(list(row.items())[:4]),
                })
                continue

            # Classe lookup
            classe_name = row.get('classe', '').strip()
            classe = classes_map.get(classe_name.lower())
            if not classe:
                errors.append({
                    'row': row_num,
                    'message': f'Classe introuvable: "{classe_name}". Vérifiez le nom exact.',
                    'data': dict(list(row.items())[:4]),
                })
                continue

            # date_of_birth validation
            dob = row.get('date_naissance') or None
            if dob:
                try:
                    from datetime import date
                    date.fromisoformat(dob)
                except ValueError:
                    errors.append({
                        'row': row_num,
                        'message': f'Date de naissance invalide "{dob}". Format attendu: YYYY-MM-DD.',
                        'data': dict(list(row.items())[:4]),
                    })
                    continue

            # Duplicate check (same name + class + dob)
            existing = Student.objects.filter(
                last_name__iexact=row['nom'],
                first_name__iexact=row['prenom'],
                classe=classe,
            ).first()
            if existing:
                skipped_count += 1
                continue

            # Create student
            try:
                student = Student(
                    last_name=row['nom'].upper(),
                    first_name=row['prenom'].capitalize(),
                    gender=genre,
                    classe=classe,
                    date_of_birth=dob or None,
                    birth_place=row.get('lieu_naissance', ''),
                    nationality=row.get('nationalite', ''),
                    parent_name=row.get('nom_parent', ''),
                    parent_phone=row.get('tel_parent', ''),
                    parent_email=row.get('email_parent', ''),
                    address=row.get('adresse', ''),
                )
                student.save()
                generate_qr_for_student(student)
                created_count += 1
            except Exception as exc:
                errors.append({
                    'row': row_num,
                    'message': str(exc),
                    'data': dict(list(row.items())[:4]),
                })

        return Response({
            'created': created_count,
            'skipped': skipped_count,
            'errors':  errors,
            'total':   created_count + skipped_count + len(errors),
        })

    @action(detail=False, methods=['get'])
    def class_summary(self, request):
        """Return all students in a class with pre-computed average and absence count."""
        classe_id = request.query_params.get('classe_id')
        if not classe_id:
            return Response({'error': 'classe_id requis'}, status=status.HTTP_400_BAD_REQUEST)

        from apps.grades.models import Grade
        from apps.attendance.models import Attendance
        from django.db.models import Avg, Count

        students = (
            Student.objects
            .filter(classe_id=classe_id, is_active=True)
            .select_related('classe', 'classe__level')
            .order_by('last_name', 'first_name')
        )

        # Grade averages (simple average of all grade values for now)
        grade_avgs = (
            Grade.objects.filter(student__classe_id=classe_id)
            .values('student_id')
            .annotate(avg=Avg('value'))
        )
        avg_map = {g['student_id']: round(float(g['avg']), 1) for g in grade_avgs}

        # Absence counts
        absent_counts = (
            Attendance.objects.filter(student__classe_id=classe_id, status='ABSENT')
            .values('student_id')
            .annotate(count=Count('id'))
        )
        absent_map = {a['student_id']: a['count'] for a in absent_counts}

        data = []
        for s in students:
            row = StudentListSerializer(s).data
            row['average'] = avg_map.get(s.id)
            row['absent_count'] = absent_map.get(s.id, 0)
            data.append(row)

        return Response(data)

    @action(detail=False, methods=['get'])
    def stats(self, request):
        total = Student.objects.filter(is_active=True).count()
        by_gender = {
            'male': Student.objects.filter(is_active=True, gender='M').count(),
            'female': Student.objects.filter(is_active=True, gender='F').count(),
        }
        by_payment = {
            'paid': Student.objects.filter(payment_status='PAID').count(),
            'pending': Student.objects.filter(payment_status='PENDING').count(),
            'overdue': Student.objects.filter(payment_status='OVERDUE').count(),
        }
        return Response({
            'total': total,
            'by_gender': by_gender,
            'by_payment': by_payment,
        })


# ── File parsing helper ────────────────────────────────────────────────────────

def _parse_file(file, ext: str) -> list[dict]:
    """Parse an uploaded .xlsx or .csv file and return a list of row dicts.

    The first row is treated as headers (case-insensitive).
    Completely empty rows are skipped.
    """
    if ext == 'xlsx':
        # Wrap in BytesIO so openpyxl always gets a seekable stream
        raw = file.read() if not isinstance(file, io.BytesIO) else file.read()
        wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)

        headers = [str(h).strip().lower() if h is not None else '' for h in next(rows_iter)]
        result = []
        for row in rows_iter:
            # Skip completely empty rows (e.g. legend row 3)
            if all(v is None or str(v).strip() == '' for v in row):
                continue
            result.append(dict(zip(headers, [str(v).strip() if v is not None else '' for v in row])))
        wb.close()
        return result

    # CSV
    content = file.read()
    # Auto-detect encoding
    try:
        text = content.decode('utf-8-sig')
    except UnicodeDecodeError:
        text = content.decode('latin-1')

    reader = csv.DictReader(io.StringIO(text))
    result = []
    for row in reader:
        if all(not v.strip() for v in row.values()):
            continue
        result.append({k.strip().lower(): v.strip() for k, v in row.items()})
    return result

